#importing modules
import pandas as pd
from openpyxl import load_workbook
pd.options.mode.chained_assignment = None

wb = load_workbook('/.../analytics-22158-credit.xlsx')
creditSheet = wb.active

#cleaing credit card data
unwantedM = " 00:00:00"
unwantedY = "2020-"
dash = '-'
#01/12/2011
#DD/MM/YYYY
#cleaning columns A, C
for i in range(273):
    if type(creditSheet.cell(row= i+1 , column = 3).value) == str:
        creditSheet.cell(row= i+1 , column = 3).value = float(creditSheet.cell(row= i+1 , column = 3).value.strip(' CR'))

    if unwantedM and unwantedY and dash in str(creditSheet.cell(row= i+1 , column = 1).value):
        creditSheet.cell(row= i+1 , column = 1).value = str(creditSheet.cell(row= i+1 , column = 1).value).replace(unwantedY,'')
        creditSheet.cell(row= i+1 , column = 1).value = str(creditSheet.cell(row= i+1 , column = 1).value).replace(unwantedM,'')
        creditSheet.cell(row= i+1 , column = 1).value = str(creditSheet.cell(row= i+1 , column = 1).value).replace(dash,'/')

#adding year 2019
for j in range(224):
    creditSheet.cell(row= j+1 , column = 1).value = str(creditSheet.cell(row= j+1 , column = 1).value) + "/2019"

#adding year 2020
for k in range(272):
    if "/2019" not in str(creditSheet.cell(row= k+1 , column = 1).value):
        creditSheet.cell(row= k+1 , column = 1).value = str(creditSheet.cell(row= k+1 , column = 1).value) + "/2020"

wb.save('/.../analytics-22158-credit.xlsx')

df = pd.read_excel('/.../analytics-22158-credit.xlsx')

def find_clean_tran(tran):
    all_clean_transaction = {
        "TOTAL FUEL 112618 CAIRO N. -07C" : "TOTAL FUEL Gas Station",
        "EL EZABY ALEX. -02G" : "EL EZABY",
        "H&M-CAIRO FESTIVAL C CAIRO N. -07C" : "H&M-CAIRO FESTIVAL",
        "AMERICAN EAGLE OUTFI CAIRO N. -07A" : "AMERICAN EAGLE OUTFIT",
        "CARREFOUR - CFC EFT CAIRO N. -07A" : "CARREFOUR",
        "MASTER EXPRESS CAIRO N. -07A" : "MASTER EXPRESS",
        "Spotify AB P0BECDAFAD Stockholm" : "Spotify",
        "EL EZABY CAIRO E. -07F" : "EL EZABY",
        "IKEA - CAIRO FESTIVAL CITY CAIRO N. -07A" : "IKEA",
        "HOLMES BURGERS ALEX. -02F" : "HOLMES BURGERS",
        "AMZN Mktp US*MA6WB3271 Amzn.com/bill" : "AMZN",
        "Go Bus HURGHADA" : "Go Bus",
        "MINISO-ALEXANDRIA CITY CENTER ALEX. -02A" : "MINISO",
        "PIZZA HUT GIZA" : "PIZZA HUT",
        "CARREFOUR EXPRESS-DOWNTOWN MALL CAIRO D.T-07D" : "CARREFOUR",
        "Uber BV Cairo" : "Uber",
        "Spotify AB P0C6599417 Stockholm" : "Spotify",
        "myfawry Maadi" : "myfawry",
        "BRAZILIAN COFFEE ALEX. -02F" : "BRAZILIAN COFFEE",
        "BERSHKA - CFC CAIRO N. -07A" : "BERSHKA",
        "DECATHLON EGYPT - CFC CAIRO E. -07C" : "DECATHLON",
        "MINISO - CFC CAIRO N. -07A" : "MINISO",
        "VAUDE 6TH OCT. -01A" : "VAUDE",
        "LC WAIKIKI - ARAB MALL 6TH OCT. -01A" : "LC WAIKIKI",
        "BOXER CAIRO E. -07E" : "BOXER",
        "TOWN TEAM TANTA -11G" : "TOWN TEAM",
        "STARBUCKS DRIVE THRU CAIRO N. -07A" : "STARBUCKS",
        "EMARAT MISR-OROOBA CAIRO N. -07A" : "EMARAT MISR",
        "UDEMY ONLINE COURSES 8888385432" : "UDEMY",
        "MCDONALD،S EMARAT MISR CAIRO N. -07A" : "MCDONALDS",
        "Swvl Transportation Cairo" : "Swvl",
        "SOUQ.COM CAIRO S. -07E" : "SOUQ",
        "EMARAT MISR - OROUBA ALEX 02G" : "EMARAT MISR",
        "CIB-MEDICARE BR BNA MEDICARE BR BNA" : "Other",
        "Spotify AB P0CDA231FA Stockholm" : "Spotify",
        "ALFA LAB CAIRO N. -07A" : "ALFA LAB",
        "AMZN Mktp US Amzn.com/bill" : "Other",
        "Spotify AB Stockholm" : "Spotify",
        "ZARA-CFC CAIRO N. -07A" : "ZARA",
        "LC WAIKIKI - CFC CAIRO N. -07A" : "LC WAIKIKI",
        "SHINY WHITE ELITE CAIRO N. -07A" : "SHINY WHITE",
        "DRINIKIES DUNES MALL 6TH OCT. -01A" : "DRINIKIES DUNES",
        "CILANTRO TAHRIR CAIRO D.T-07D" : "CILANTRO",
        "NBE ATM360 CAIRO" : "Other",
        "EL AMIN CAIRO N. -07A" : "EL AMIN",
        "ADIDAS -CFC CAIRO N. -07A" : "ADIDAS",
        "ON THE RUN -URBAN 5 NE CAIRO" : "ON THE RUN",
        "MAHRAGA REST CAIRO" : "MAHRAGA",
        "HEART ATTACK CAIRO S. -07E" : "HEART ATTACK",
        "SPINNEYS 7 CAIRO" : "SPINNEYS",
        "CIB-CFC BR. II CFC BR. II" : "Other",
        "CAREEM MO 6TH OCT. -01A" : "CAREEM",
        "MCDONALD S SHELL 90 CAIRO N. -07A" : "MCDONALDS",
        "TOTAL BONJOUR 502894 CAIRO N. -07A" : "TOTAL BONJOUR Gas Station",
        "BEST BUY - SAN STEFANO ALEX" : "Other",
        "EGYPT RAILWAYS PUBLIC RAMSES" : "EGYPT RAILWAYS",
        "CIB-RAMSIS BR 2 RAMSIS BR 2" : "Other",
        "EL HAMMADY CAR SERVICES-MOBIL ALEX. -02G" : "TOTAL FUEL Gas Station",
        "CIB-ROUSHDY BR ROUSHDY BR" : "Other",
        "KFC - MEHY EL DINE GIZA -12E" : "KFC",
        "CIB-DOWN TOWN_MALL DOWN_TOWN_MALL" : "Other"
    }

    clean_transaction = all_clean_transaction[tran]

    return clean_transaction


for i in range(len(df['Transaction'])):
    df['Transaction'].loc[i] = find_clean_tran(df['Transaction'].loc[i])


df.to_csv(r'/.../analytics-22158-credit.csv', index = False)

#cleaing debit card data
df = pd.read_csv('/.../analytics-22158-debit.csv')
def get_date(date):
    date_by_day = date[0:2]
    search_key = date[2:5]
    
    months = {
        # 27/06/2019
        # 02Jun19
        "Jun" : "/06/2019",
        "Jul" : "/07/2019",
        "Aug" : "/08/2019",
        "Sep" : "/09/2019",
        "Oct" : "/10/2019",
        "Nov" : "/11/2019",
        "Dec" : "/12/2019",
        "Jan" : "/01/2020",
        "Feb" : "/02/2020",
        "Mar" : "/03/2020"
    }
    for key, val in months.items():
            if search_key in key:
                clean_date = date_by_day + val
                return clean_date
    # res = [val for key, val in months.items() search_key for search_key in search_keys if search_key in key]
    # print(res)

def find_clean_tran(tran):
    transaction_lst = [
        "PAPA JOHNS",
        "Uber",
        "Spotify",
        "WITHDRAWAL",
        "myfawry",
        "PIZZAHUT",
        "Swvl",
        "Other",
        "EMARATMISR",
        "Go Bus",
        "STARBUCKS",
        "HEART ATTACK",
        "DUKES",
        "IKEA",
        "CARREFOUR",
        "KFC",
        "RADIOSHACK",
        "CILANTRO",
        "EGYPTRAILWAYS",
        "MCDONALD"
    ]
    for i in range(len(transaction_lst)):
        if transaction_lst[i] in tran:
            return transaction_lst[i]

idx = df[df['Transaction'].str.contains('SALARY|CENTRIS|Internet Transfer|OPENING|CLOSING|DEPOSIT|5264390000004804|08520600', na=False)]
idx_list = list(idx.index)
# df = df.drop(df.index[idx_list])
df = df.dropna()

for i in range(len(df)):
    #df['Transaction'].loc[i] = find_clean_tran(df['Transaction'].loc[i])
    print(df['Date'].loc[i])
    #df['Date'].loc[i] = get_date(df['Date'].loc[i])

df.to_csv(r'/.../analytics-22158-debit.csv', index = False)